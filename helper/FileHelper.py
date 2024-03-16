from openpyxl import Workbook
import os


class FileHelper:

    def __init__(self):
        self._output_folder = "reports"

    def write_to_excel_suspended_licenses(self, licenses):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Nume", "Prenume", "Categorie", "Data De Emitere", "Data de Expirare", "Suspendat"])
        for license in licenses:
            ws.append([
                license.id,
                license.nume,
                license.prenume,
                license.categorie,
                license.dataDeEmitere,
                license.dataDeExpirare,
                license.suspendat
            ])
        output_path = os.path.join(self._output_folder, "suspended_licenses.xlsx")
        try:
            os.makedirs(self._output_folder)
        except FileExistsError as e:
            pass
        wb.save(output_path)

    def write_to_excel_valid_licenses(self, licenses):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Nume", "Prenume", "Categorie", "Data De Emitere", "Data de Expirare", "Suspendat"])
        for license in licenses:
            ws.append([
                license.id,
                license.nume,
                license.prenume,
                license.categorie,
                license.dataDeEmitere,
                license.dataDeExpirare,
                license.suspendat
            ])
        output_path = os.path.join(self._output_folder, "valid_licenses.xlsx")
        try:
            os.makedirs(self._output_folder)
        except FileExistsError as e:
            pass
        wb.save(output_path)

    def write_to_excel_by_category_licenses(self, licenses):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Nume", "Prenume", "Categorie", "Data De Emitere", "Data de Expirare", "Suspendat"])
        for license in licenses:
            ws.append([
                license.id,
                license.nume,
                license.prenume,
                license.categorie,
                license.dataDeEmitere,
                license.dataDeExpirare,
                license.suspendat
            ])
        output_path = os.path.join(self._output_folder, "licenses_by_category.xlsx")
        try:
            os.makedirs(self._output_folder)
        except FileExistsError as e:
            pass
        wb.save(output_path)